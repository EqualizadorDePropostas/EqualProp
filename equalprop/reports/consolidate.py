import csv
import re
import openpyxl
import os
import shutil
import time
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

from .suppliers import generate_suppliers_report
from .globals import generate_preco_report
from .comparison import generate_comparison_report
from .relatorio_rfp_cabecalho import generate_rfp_header_report
from .relatorio_condicomer import generate_condicomer_report
from .relatorio_socios import generate_socios_report


def consolidate_reports(rfp_json=None, propostas_json=None, condicomer_padronizadas=None, quadros_societarios=None, socio_comum=None, out_dir=None):
    """Gera e consolida relatorios em CSV e Excel.

    Ordem:
    1) relatorio_rfp_cabecalho.csv
    2) relatorio_fornecedores.csv
    3) relatorio_preco.csv
    4) relatorio_condicomer.csv
    5) relatorio_socios.csv
    6) comparacao_produtos.csv
    """
    # Preparar out_dir e entrar nele para nao misturar com execucoes anteriores
    if out_dir is None:
        out_dir = os.path.abspath('out')
    else:
        out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    try:
        for name in os.listdir(out_dir):
            path = os.path.join(out_dir, name)
            try:
                if os.path.isfile(path) or os.path.islink(path):
                    os.unlink(path)
                elif os.path.isdir(path):
                    shutil.rmtree(path)
            except Exception:
                pass
    except Exception:
        pass

    start_ts = time.time()
    # Gerar arquivos, se possivel
    generated_files = {}

    def _safe_generate(generator, filename, *args, **kwargs):
        target = os.path.abspath(filename)
        try:
            generator(*args, filename=target, **kwargs)
            return target
        except PermissionError:
            ts = time.strftime('%Y%m%d_%H%M%S')
            base, ext = os.path.splitext(target)
            fallback = f"{base}__{ts}{ext}"
            print(f"[AVISO] {os.path.basename(target)} em uso. Salvando em {fallback}")
            try:
                generator(*args, filename=fallback, **kwargs)
                return fallback
            except Exception as inner:
                print(f"[ERRO] Falha ao salvar fallback para {os.path.basename(target)}: {inner}")
        except Exception as e:
            print(f"[AVISO] Falha ao gerar {os.path.basename(target)}: {e}")
        return None

    if rfp_json is not None:
        path_rfp = _safe_generate(
            generate_rfp_header_report,
            os.path.join(out_dir, 'relatorio_rfp_cabecalho.csv'),
            rfp_json
        )
        if path_rfp:
            generated_files['relatorio_rfp_cabecalho.csv'] = path_rfp

    if propostas_json is not None:
        path_suppliers = _safe_generate(
            generate_suppliers_report,
            os.path.join(out_dir, 'relatorio_fornecedores.csv'),
            propostas_json
        )
        if path_suppliers:
            generated_files['relatorio_fornecedores.csv'] = path_suppliers

    if rfp_json is not None and propostas_json is not None:
        path_preco = _safe_generate(
            generate_preco_report,
            os.path.join(out_dir, 'relatorio_preco.csv'),
            rfp_json,
            propostas_json
        )
        if path_preco:
            generated_files['relatorio_preco.csv'] = path_preco

        path_comparacao = _safe_generate(
            generate_comparison_report,
            os.path.join(out_dir, 'comparacao_produtos.csv'),
            rfp_json,
            propostas_json
        )
        if path_comparacao:
            generated_files['comparacao_produtos.csv'] = path_comparacao

    if condicomer_padronizadas is not None:
        path_condicomer = _safe_generate(
            generate_condicomer_report,
            os.path.join(out_dir, 'relatorio_condicomer.csv'),
            condicomer_padronizadas
        )
        if path_condicomer:
            generated_files['relatorio_condicomer.csv'] = path_condicomer
    # Gerar relatorio_socios.csv independentemente dos dados disponiveis
    path_socios = None
    cwd_before = os.getcwd()
    try:
        os.chdir(out_dir)
        result_path = generate_socios_report(quadros_societarios or {}, socio_comum or {})
        if not result_path:
            result_path = 'relatorio_socios.csv'
        if not os.path.isabs(result_path):
            result_path = os.path.join(out_dir, result_path)
        path_socios = os.path.abspath(result_path)
    except PermissionError as e:
        print(f"[AVISO] Falha ao gerar relatorio_socios.csv: {e}")
    except Exception as e:
        print(f"[AVISO] Falha ao gerar relatorio_socios.csv: {e}")
    finally:
        try:
            os.chdir(cwd_before)
        except Exception:
            pass
    fallback_socios_path = os.path.join(out_dir, 'relatorio_socios.csv')
    if path_socios and os.path.exists(path_socios):
        generated_files['relatorio_socios.csv'] = path_socios
    else:
        try:
            with open(fallback_socios_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Quadro de socios e administradores', '', '', '', ''])
                writer.writerow(['Propostas em que ha socios em comum ', '', '', '', ''])
        except Exception as e:
            print(f"[AVISO] Falha ao criar fallback de relatorio_socios.csv: {e}")
        else:
            generated_files['relatorio_socios.csv'] = fallback_socios_path


    for logical_name, actual_path in list(generated_files.items()):
        canonical_path = os.path.join(out_dir, logical_name)
        if os.path.abspath(actual_path) != os.path.abspath(canonical_path):
            try:
                shutil.copyfile(actual_path, canonical_path)
                generated_files[logical_name] = canonical_path
            except PermissionError:
                print(f"[AVISO] {logical_name} esta em uso. Nova versao salva em {actual_path}")
            except Exception as copy_err:
                print(f"[AVISO] Nao foi possivel atualizar {logical_name}: {copy_err}")

    # Substituir esta seção (linhas ~60-80) por:
    for nome in (
        "relatorio_rfp_cabecalho.csv",
        "relatorio_fornecedores.csv",
        "relatorio_preco.csv",
        "relatorio_socios.csv",
        "comparacao_produtos.csv",
    ):
        src = os.path.abspath(nome)
        dst = os.path.join(out_dir, nome)
        try:
            if os.path.exists(src):
                # SEM verificação de timestamp - sempre mover se existir
                if not os.path.exists(dst):
                    shutil.move(src, dst)
        except Exception:
            pass

    # NÃO mover relatorio_condicomer.csv pois já foi gerado direto no out_dir

    
    tabela_specs = [
        ('relatorio_rfp_cabecalho.csv', os.path.join(out_dir, 'relatorio_rfp_cabecalho.csv')),
        ('relatorio_fornecedores.csv', os.path.join(out_dir, 'relatorio_fornecedores.csv')),
        ('relatorio_preco.csv', os.path.join(out_dir, 'relatorio_preco.csv')),
        ('relatorio_condicomer.csv', os.path.join(out_dir, 'relatorio_condicomer.csv')),
        ('relatorio_socios.csv', os.path.join(out_dir, 'relatorio_socios.csv')),
        ('comparacao_produtos.csv', os.path.join(out_dir, 'comparacao_produtos.csv')),
    ]
    tabelas = []
    for logical_name, default_path in tabela_specs:
        tabelas.append((logical_name, generated_files.get(logical_name, default_path)))
    arquivo_csv = "relatorio_consolidado.csv"
    arquivo_xlsx = "relatorio_consolidado.xlsx"

    linhas_consolidadas = []
    tabela_dados = []
    total_tabelas = len(tabelas)
    for idx, (logical_name, file_path) in enumerate(tabelas):
        rows = []
        success = False
        try:
            with open(file_path, 'r', encoding='utf-8') as arquivo:
                rows = list(csv.reader(arquivo))
            success = True
        except FileNotFoundError:
            print(f"[AVISO] Arquivo nao encontrado: {logical_name}")
        except PermissionError:
            print(f"[AVISO] Nao foi possivel ler {logical_name}: arquivo em uso")
        if success:
            linhas_consolidadas.extend(rows)
            trailing_blank = idx < total_tabelas - 1
            tabela_dados.append({
                'name': logical_name,
                'rows': rows,
                'first_cols': len(rows[0]) if rows else 0,
                'trailing_blank': trailing_blank
            })
            if trailing_blank:
                linhas_consolidadas.append([])  # 1 linha em branco

    # Inserir uma linha vazia no topo (acima do relatorio_rfp_cabecalho)
    linhas_consolidadas.insert(0, [])

    csv_path = os.path.join(out_dir, arquivo_csv)
    with open(csv_path, 'w', newline='', encoding='utf-8') as arquivo:
        csv.writer(arquivo).writerows(linhas_consolidadas)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in linhas_consolidadas:
        sheet.append(row)

    # Mapear secoes (inicio, quantidade de linhas e colunas da primeira linha)
    secoes = []
    _current = 2
    for data in tabela_dados:
        rows = data['rows']
        _start = _current
        secoes.append({
            'name': data['name'],
            'start': _start,
            'rows': len(rows),
            'first_cols': data['first_cols']
        })
        _current += len(rows)
        if data.get('trailing_blank'):
            _current += 1  # linha em branco de separacao

    # 1) Substituir 'excel formula 1' pelo RESULTADO (nao formula)
    def _is_null(x):
        return isinstance(x, str) and x.strip().lower() == 'null'

    def _is_empty(x):
        return x is None or (isinstance(x, str) and x == '')

    def _parse_num(x):
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        if isinstance(x, str):
            s = x.strip()
            if s == '' or s.lower() == 'null':
                return None
            s = s.replace(' ', '')
            s = s.replace(',', '.')
            try:
                return float(s)
            except Exception:
                return None
        return None

    def _fmt_comma(v: float) -> str:
        s = f"{v:.3f}"
        if '.' in s:
            s = s.rstrip('0').rstrip('.')
        return s.replace('.', ',')

    def _number_format_for(value) -> str:
        try:
            v = float(value)
        except (TypeError, ValueError):
            return '#,##0.00'
        if abs(v) >= 100:
            return '#,##0'
        return '#,##0.00'

    # 1.a) excel formula 3 -> minimo entre todas as colunas 'R$ unit' de propostas na mesma linha.
    # Detecta dinamicamente as colunas de preco unitario como todos os indices 6,9,12,...
    # imediatamente antes da propria celula 'excel formula 3'.
    for r_idx, row in enumerate(linhas_consolidadas, start=1):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, str) and val.strip() and val.strip().lower().replace(' ', '') == 'excelformula3':
                nums = []
                # colunas 6, 9, 12, ... ate a coluna anterior ao 'excel formula 3'
                for col in range(6, c_idx, 3):
                    v = sheet.cell(row=r_idx, column=col).value
                    if _is_null(v) or _is_empty(v):
                        continue
                    pv = _parse_num(v)
                    if pv is not None:
                        nums.append(pv)
                target = sheet.cell(row=r_idx, column=c_idx)
                if nums:
                    target.value = _fmt_comma(min(nums))
                else:
                    target.value = ''

    # 1.b) excel formula 1 -> result (left cell multiplied by quantity in column D)
    for r_idx, row in enumerate(linhas_consolidadas, start=1):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, str) and val.strip() and val.strip().lower().replace(' ', '') == 'excelformula1' and c_idx > 1:
                left_val = sheet.cell(row=r_idx, column=c_idx - 1).value
                d_val = sheet.cell(row=r_idx, column=4).value
                target = sheet.cell(row=r_idx, column=c_idx)
                if _is_null(left_val) or _is_null(d_val):
                    target.value = 'null'
                    continue
                if _is_empty(left_val) and _is_empty(d_val):
                    target.value = ''
                    continue
                a = _parse_num(left_val)
                b = _parse_num(d_val)
                if a is None or b is None:
                    target.value = ''
                else:
                    try:
                        target.value = float(a * b)
                    except Exception:
                        target.value = a * b
                    target.number_format = _number_format_for(target.value)

    # 1.c) excel formula 2 -> sum of the 'R$ TOTAL' above in the same column (same proposal)
    def _row_is_blank(r: int) -> bool:
        for col in range(1, sheet.max_column + 1):
            v = sheet.cell(row=r, column=col).value
            if not _is_empty(v):
                return False
        return True

    for r_idx, row in enumerate(linhas_consolidadas, start=1):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, str) and val.strip() and val.strip().lower().replace(' ', '') == 'excelformula2':
                has_null_above = False
                has_nonempty_above = False
                total = 0.0
                has_number = False
                rr = r_idx - 1
                # Go upwards until a blank separator row
                while rr >= 1 and not _row_is_blank(rr):
                    v = sheet.cell(row=rr, column=c_idx).value
                    if _is_null(v):
                        has_null_above = True
                    if not _is_empty(v):
                        has_nonempty_above = True
                    pv = _parse_num(v)
                    if pv is not None:
                        total += pv
                        has_number = True
                    rr -= 1
                target = sheet.cell(row=r_idx, column=c_idx)
                if has_null_above:
                    target.value = 'null'
                elif not has_nonempty_above:
                    target.value = ''
                elif has_number:
                    try:
                        target.value = float(total)
                    except Exception:
                        target.value = total
                    target.number_format = _number_format_for(target.value)
                else:
                    target.value = ''

    # 3) Aplicar bordas verticais direitas (sera feito apos mapear secoes)

    # 4) Congelar as colunas A..E (pane em F1)
    sheet.freeze_panes = sheet['F1']

    # 5) Definir fonte tamanho 10 para toda a planilha
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = Font(size=10)

    # 6) Aplicar negrito e azul marinho (navy) conforme solicitado
    bold_navy = Font(size=10, bold=True, color="1E40FF")

    def _style_header_row(sec):
        if sec['rows'] <= 0:
            return
        r = sec['start']
        ncols = sec.get('first_cols') or 0
        if ncols <= 0:
            ncols = sheet.max_column
        for c in range(1, ncols + 1):
            sheet.cell(row=r, column=c).font = bold_navy

    def _style_first_col(sec):
        if sec['rows'] <= 0:
            return
        for rr in range(sec['start'], sec['start'] + sec['rows']):
            sheet.cell(row=rr, column=1).font = bold_navy

    by_name = {s['name']: s for s in secoes}

    # Aplicar bordas verticais direitas (dinamicas conforme numero de propostas)
    thick_black = Side(border_style='thick', color='000000')
    thin_white = Side(border_style='thin', color='FFFFFF')

    black_cols = {5}
    preco_sec = by_name.get('relatorio_preco.csv')
    if preco_sec and preco_sec.get('rows', 0) >= 1:
        # Cabecalho do relatorio_preco e sempre de uma unica linha
        hdr_row = preco_sec['start']
        c = 6
        while c + 2 <= sheet.max_column:
            v1 = sheet.cell(row=hdr_row, column=c).value
            v2 = sheet.cell(row=hdr_row, column=c + 1).value
            v3 = sheet.cell(row=hdr_row, column=c + 2).value
            if isinstance(v1, str) and isinstance(v2, str) and isinstance(v3, str):
                if v1.strip().lower() == 'r$ unit' and v2.strip().lower() == 'r$ total' and v3.strip().lower().startswith('semelhan'):
                    black_cols.add(c + 2)
                    c += 3
                    continue
            break
        # Nenhum fallback necessario; so existe uma linha de cabecalho
        sep_col = c + 2
        if sep_col <= sheet.max_column:
            black_cols.add(sep_col)
    socios_sec = by_name.get('relatorio_socios.csv')
    if socios_sec:
        ncols_socios = socios_sec.get('first_cols') or sheet.max_column
        c = 6
        while c + 2 <= ncols_socios:
            black_cols.add(c + 2)
            c += 3


    max_row = sheet.max_row
    max_col = sheet.max_column
    white_cols = set(range(1, max_col + 1)) - black_cols
    for r in range(1, max_row + 1):
        for c in black_cols:
            cell = sheet.cell(row=r, column=c)
            cell.border = Border(right=thick_black)
        for c in white_cols:
            cell = sheet.cell(row=r, column=c)
            cell.border = Border(right=thin_white)

    # primeira coluna do relatorio_rfp_cabecalho
    sec = by_name.get('relatorio_rfp_cabecalho.csv')
    if sec:
        _style_first_col(sec)

    # Primeira linha e primeira coluna do relatorio_fornecedores
    sec = by_name.get('relatorio_fornecedores.csv')
    if sec:
        _style_header_row(sec)
        _style_first_col(sec)

    # Primeira linha do relatorio_preco
    sec = by_name.get('relatorio_preco.csv')
    if sec:
        # Cabecalho (uma linha): aplicar negrito e azul
        r1 = sec['start']
        ncols = sec.get('first_cols') or sheet.max_column
        for c in range(1, ncols + 1):
            sheet.cell(row=r1, column=c).font = bold_navy

        # Alturas e alinhamento do corpo
        # Se existe apenas uma linha de cabecalho (r2==r1), o corpo comeca em r2+1
        start_body = r1 + 1
        end_body = sec['start'] + sec['rows'] - 1
        # Linhas de dados (exceto a ultima "Total"): 65 pt e alinhamento a esquerda e vertical superior
        for rr in range(start_body, max(start_body, end_body)):
            sheet.row_dimensions[rr].height = 65
            for c in range(1, ncols + 1):
                sheet.cell(row=rr, column=c).alignment = Alignment(horizontal='left', vertical='top')
        # ultima linha (Total), se existir: manter altura padrao mas alinhar tambem a esquerda/topo
        if end_body >= start_body:
            for c in range(1, ncols + 1):
                sheet.cell(row=end_body, column=c).alignment = Alignment(horizontal='left', vertical='top')
        # Quebra de linha na coluna Descricao (coluna C)
        for rr in range(start_body, end_body + 1):
            cell = sheet.cell(row=rr, column=3)
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
            # Remover introducoes como "descricao do produto:" se presentes
            if isinstance(cell.value, str):
                try:
                    cell.value = re.sub(r'(?i)descri[cc][aa]o\s*(do\s*produto)?\s*[:\-]\s*', '', cell.value)
                except Exception:
                    pass

        # Formatacao das colunas "R$ unit" e "R$ total"
        def _fmt_mixed(v: float) -> str:
            if v < 100:
                s = f"{v:.2f}"
            else:
                s = f"{v:.0f}"
            return s.replace('.', ',')

        header_row = r1  # unica linha do cabecalho
        unit_cols = []
        total_cols = []
        for c in range(1, sheet.max_column + 1):
            hv = sheet.cell(row=header_row, column=c).value
            if isinstance(hv, str):
                hs = hv.strip().lower()
                if hs == 'r$ unit':
                    unit_cols.append(c)
                elif hs == 'r$ total':
                    total_cols.append(c)

        price_cols = set(unit_cols + total_cols)
        for rr in range(start_body, end_body + 1):
            # Formatar quantidade (coluna D)
            try:
                qcell = sheet.cell(row=rr, column=4)
                qv = qcell.value
                if not (isinstance(qv, str) and qv.strip().lower() in ('', 'null')):
                    qpv = _parse_num(qv)
                    if qpv is not None:
                        try:
                            qcell.value = float(qpv)
                        except Exception:
                            qcell.value = qpv
                        qcell.number_format = _number_format_for(qcell.value)
            except Exception:
                pass
            for c in price_cols:
                cell = sheet.cell(row=rr, column=c)
                v = cell.value
                # pular 'null' e vazios
                if isinstance(v, str) and v.strip().lower() in ('', 'null'):
                    continue
                pv = _parse_num(v)
                if pv is not None:
                    try:
                        cell.value = float(pv)
                    except Exception:
                        cell.value = pv
                    cell.number_format = _number_format_for(cell.value)
                    

    # Primeira linha do relatorio_condicomer
    sec = by_name.get('relatorio_condicomer.csv')
    if sec:
        _style_header_row(sec)
        # Linhas (exceto cabecalho) com altura 65pt, wrap e alinhamento esquerda/topo
        start_body = sec['start'] + 1
        end_body = sec['start'] + sec['rows'] - 1
        ncols = sec.get('first_cols') or sheet.max_column
        for rr in range(start_body, end_body + 1):
            sheet.row_dimensions[rr].height = 65
            for c in range(1, ncols + 1):
                cell = sheet.cell(row=rr, column=c)
                # Sem normalizacao artificial de quebras; manter conteudo original
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Merge das 3 colunas de cada proposta (valor + 2 vazias), mantendo o conteudo da celula da esquerda
        start_merge = sec['start']
        end_merge = sec['start'] + sec['rows'] - 1
        ncols_merge = sec.get('first_cols') or sheet.max_column
        for rr in range(start_merge, end_merge + 1):
            c = 6
            while c + 2 <= ncols_merge:
                sheet.merge_cells(start_row=rr, start_column=c, end_row=rr, end_column=c + 2)
                # Borda direita preta e grossa no limite direito da trinca
                right_cell = sheet.cell(row=rr, column=c + 2)
                b = right_cell.border
                right_cell.border = Border(left=b.left, right=Side(border_style='thick', color='000000'), top=b.top, bottom=b.bottom)
                c += 3

    # Linhas do relatorio_socios
    sec = by_name.get('relatorio_socios.csv')
    if sec:
        _style_header_row(sec)
        ncols = sec.get('first_cols') or sheet.max_column
        start_row = sec['start']
        end_row = sec['start'] + sec['rows'] - 1
        left_cols = []
        c = 6
        while c + 2 <= ncols:
            left_cols.append(c)
            c += 3
        if left_cols:
            wrap_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
            data_start = start_row + 2
            for rr in range(start_row, end_row + 1):
                if rr >= data_start:
                    current_height = sheet.row_dimensions[rr].height
                    sheet.row_dimensions[rr].height = current_height * 2 if current_height else 30
                for left_col in left_cols:
                    sheet.merge_cells(start_row=rr, start_column=left_col, end_row=rr, end_column=left_col + 2)
                    cell = sheet.cell(row=rr, column=left_col)
                    if rr >= data_start:
                        if cell.value is None:
                            cell.value = ''
                        cell.alignment = wrap_top_left
            for rr in range(start_row, end_row + 1):
                for left_col in left_cols:
                    thick_target = sheet.cell(row=rr, column=left_col + 2)
                    b = thick_target.border
                    thick_target.border = Border(left=b.left, right=Side(border_style='thick', color='000000'), top=b.top, bottom=b.bottom)
                    left_cell = sheet.cell(row=rr, column=left_col)
                    bleft = left_cell.border
                    left_cell.border = Border(left=bleft.left, right=Side(border_style='thin', color='FFFFFF'), top=bleft.top, bottom=bleft.bottom)
            second_row = start_row + 1
            if second_row <= sheet.max_row:
                red_fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
                for left_col in left_cols:
                    cell = sheet.cell(row=second_row, column=left_col)
                    value = cell.value
                    if isinstance(value, str) and value.strip().lower() == 'sim':
                        cell.fill = red_fill

    # Primeira linha do comparacao_produtos
    sec = by_name.get('comparacao_produtos.csv')
    if sec:
        _style_header_row(sec)

    # 7) Largura da coluna C ~320px (aprox width=45)
    sheet.column_dimensions['C'].width = 45

    # Linhas separadoras (1 linha acima de cada tabela): fundo cinza claro e bordas laterais cinza
    gray = 'D9D9D9'
    gray_fill = PatternFill(fill_type='solid', start_color=gray, end_color=gray)
    gray_side = Side(border_style='thin', color=gray)
    for sec in secoes:
        sep_row = sec['start'] - 1
        if sep_row < 1:
            continue
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=sep_row, column=c)
            cell.fill = gray_fill
            b = cell.border
            cell.border = Border(left=gray_side, right=gray_side, top=b.top, bottom=b.bottom)

    xlsx_path = os.path.join(out_dir, arquivo_xlsx)
    workbook.save(xlsx_path)

    abs_csv = os.path.abspath(csv_path)
    abs_xlsx = os.path.abspath(xlsx_path)
    print(f"[OK] Relatorio consolidado: {abs_csv} e {abs_xlsx}")
    return abs_csv, abs_xlsx


























